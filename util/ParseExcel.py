from openpyxl import load_workbook
from config.varCondig import *
class ParseExcel(object):
    '''
    解析excel文件的封装
    '''
    def __init__(self):
        # 加载excel文件到内存
        self.wb = load_workbook(testExcelValuePath)

    def getRowValue(self, sheetName, rawNo):
        '''
        获取某一行的数据
        :param sheetName:
        :param rawNo:
        :return: 列表
        '''
        sh = self.wb[sheetName]
        rowValueList = []
        for y in range(2, sh.max_column+1):
            value = sh.cell(rawNo,y).value
            rowValueList.append(value)
        return rowValueList
    def getColumnValue(self, sheetName, colNo):
        '''
        获取某一列的数据
        :param sheetName:
        :param colNo:
        :return: 列表
        '''
        sh = self.wb[sheetName]
        colValueList = []
        for x in range(2, sh.max_row +1):
            value = sh.cell(x, colNo).value
            colValueList.append(value)
        return colValueList

    def getCellOfValue(self, sheetName, rowNo, colNo):
        '''
        获取某一个单元格的数据
        :param sheetName:
        :param rowNo:
        :param colNo:
        :return: 字符串
        '''
        sh = self.wb[sheetName]
        value = sh.cell(rowNo, colNo).value
        return value
    def writeCell(self, sheetName, rowNo, colNo, value):
        '''
        向某个单元格写入数据
        :param rowNo: 行号
        :param colNo: 列号
        :param value:
        :return: 无
        '''
        sh = self.wb[sheetName]
        sh.cell(rowNo, colNo).value = value
        self.wb.save(testExcelValuePath)
if __name__=='__main__':
    p = ParseExcel()
    print(p.getRowValue('126account',2))
    print(p.getColumnValue('126account',3))
    print(p.getCellOfValue('126account', 2, 3))
